#!/usr/bin/env python3
"""
Build glossary.json and glossary_processed.json from an Excel workbook.

Excel columns (case-insensitive):
- Terms
- Technical
- Layman

Outputs:
- glossary.json            (raw; [[Term]] -> Term)
- glossary_processed.json  (processed; [[Term]] -> lightweight span with data-id for JS tooltip)

Usage:
  python build_glossary_from_excel.py /path/to/glossary.xlsx /path/to/output_dir

Dependencies:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import html
import json
import os
import re
from dataclasses import dataclass
from typing import Dict, List

import openpyxl

BRACKET_REF_RE = re.compile(r"\[\[([^\]|]+?)(?:\|([^\]]+?))?\]\]")  # [[target]] or [[target|display]]


def slugify(term: str) -> str:
    s = term.strip().lower()
    s = s.replace("&", " and ")
    s = s.replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "term"


def make_unique_id(base_id: str, used: Dict[str, int]) -> str:
    if base_id not in used:
        used[base_id] = 1
        return base_id
    used[base_id] += 1
    return f"{base_id}-{used[base_id]}"


def normalize_term_key(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def replace_refs_raw(text: str) -> str:
    """[[target]] -> target, [[target|display]] -> display"""
    if not text:
        return ""

    def repl(m: re.Match) -> str:
        target = (m.group(1) or "").strip()
        display = (m.group(2) or "").strip()
        return display if display else target

    return BRACKET_REF_RE.sub(repl, text)


@dataclass
class Entry:
    id: str
    term: str
    technical: str
    layman: str


def read_excel_entries(xlsx_path: str) -> List[Entry]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    needed = {"terms", "technical", "layman"}
    chosen_ws = None
    headers = None
    start_row = None

    for ws in wb.worksheets:
        header_row = None
        for r in range(1, min(ws.max_row, 20) + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if any(v is not None and str(v).strip() for v in row_vals):
                header_row = r
                break
        if header_row is None:
            continue

        hdr_map: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if v is None:
                continue
            key = str(v).strip().lower()
            if key in needed:
                hdr_map[key] = c

        if needed.issubset(hdr_map.keys()):
            chosen_ws = ws
            headers = hdr_map
            start_row = header_row + 1
            break

    if chosen_ws is None or headers is None or start_row is None:
        raise ValueError("Could not find columns named Terms, Technical, Layman (case-insensitive).")

    used_ids: Dict[str, int] = {}
    entries: List[Entry] = []

    for r in range(start_row, chosen_ws.max_row + 1):
        term_val = chosen_ws.cell(row=r, column=headers["terms"]).value
        if term_val is None or not str(term_val).strip():
            continue

        term = str(term_val).strip()
        technical_val = chosen_ws.cell(row=r, column=headers["technical"]).value
        layman_val = chosen_ws.cell(row=r, column=headers["layman"]).value

        technical = ("" if technical_val is None else str(technical_val)).strip()
        layman = ("" if layman_val is None else str(layman_val)).strip()

        unique_id = make_unique_id(slugify(term), used_ids)
        entries.append(Entry(id=unique_id, term=term, technical=technical, layman=layman))

    entries.sort(key=lambda e: e.term.lower())
    return entries


def build_lookup(entries: List[Entry]) -> Dict[str, Entry]:
    """
    Lookup by normalized term string.

    Also adds aliases for parenthetical abbreviations, e.g.
    "Command and Control (C2)" -> lookup works for both:
      - "Command and Control (C2)"
      - "C2"
    """
    lut: Dict[str, Entry] = {}

    paren_alias_re = re.compile(r"^(.*?)\s*\(([^()]+)\)\s*$")

    for e in entries:
        full_key = normalize_term_key(e.term)
        lut[full_key] = e

        match = paren_alias_re.match(e.term.strip())
        if match:
            base_term = match.group(1).strip()
            alias = match.group(2).strip()

            if base_term:
                lut.setdefault(normalize_term_key(base_term), e)
            if alias:
                lut.setdefault(normalize_term_key(alias), e)

    return lut


def replace_refs_processed(text: str, lut: Dict[str, Entry]) -> str:
    """
    [[target]] -> <span class="glossary-term--ref" data-id="id">target</span>
    [[target|display]] -> <span class="glossary-term--ref" data-id="id">display</span>
    """
    if not text:
        return ""

    def repl(m: re.Match) -> str:
        target = (m.group(1) or "").strip()
        display = (m.group(2) or "").strip()

        key = normalize_term_key(target)
        visible = display if display else target
        visible_esc = html.escape(visible)

        entry = lut.get(key)
        if entry is None:
            return (
                f'<span class="glossary-term--ref glossary-term--missing" data-id="">'
                f"{visible_esc}</span>"
            )

        return (
            f'<span class="glossary-term--ref" '
            f'data-id="{html.escape(entry.id)}">'
            f"{visible_esc}</span>"
        )

    return BRACKET_REF_RE.sub(repl, text)


def entries_to_json(entries: List[Entry], *, processed: bool, lut: Dict[str, Entry]) -> List[dict]:
    out: List[dict] = []
    for e in entries:
        if processed:
            technical = replace_refs_processed(e.technical, lut)
            layman = replace_refs_processed(e.layman, lut)
        else:
            technical = replace_refs_raw(e.technical)
            layman = replace_refs_raw(e.layman)

        out.append(
            {
                "id": e.id,
                "term": e.term,
                "technical": technical,
                "layman": layman,
            }
        )
    return out


def write_json_pretty(path: str, data: List[dict]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, separators=(",", ": "))
        f.write("\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build glossary JSON files from Excel.")
    parser.add_argument("input_excel", help="Path to the input .xlsx file")
    parser.add_argument("output_dir", help="Directory to write glossary.json files")
    args = parser.parse_args()

    entries = read_excel_entries(args.input_excel)
    lut = build_lookup(entries)

    raw = entries_to_json(entries, processed=False, lut=lut)
    processed = entries_to_json(entries, processed=True, lut=lut)

    write_json_pretty(os.path.join(args.output_dir, "glossary.json"), raw)
    write_json_pretty(os.path.join(args.output_dir, "glossary_processed.json"), processed)

    print("Wrote glossary.json and glossary_processed.json")


if __name__ == "__main__":
    main()